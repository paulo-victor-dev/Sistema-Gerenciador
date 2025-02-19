import sqlite3
import openpyxl.workbook
import pandas as pd

class Acao_SQL: 
    def acao_sql(self, query, parametro=False, fetchall=False):
        '''Abre a conexão com o banco de dados, executa uma query SQL, fecha a conexão e retorna o resultado da query.'''
        caminho_banco = r'Base de dados\Base de Dados.db'

        try:
            self.conexao = sqlite3.connect(caminho_banco)

            self.cursor = self.conexao.cursor()

            if parametro:
                self.cursor.execute(query, parametro)
            else:
                self.cursor.execute(query)

            if fetchall:
                itens = self.cursor.fetchall()

            self.conexao.commit()
        
        except sqlite3.IntegrityError as e:
            erro_integridade = 'ERRO INTEGRIDADE'
            return erro_integridade
        
        except Exception as e:
            print(f'Algo deu errado: {e}')

        finally:
            self.cursor.close()
            self.conexao.close()

        return itens if fetchall else None

class Cadastrar(Acao_SQL):
    def cadastrar_pedidos(self):
        pass

    def cadastrar_itens_do_pedido(self): 
        pass

    def cadastrar_clientes(self, nome, cpf, email, telefone):
        resultado_acao = self.acao_sql('''
        INSERT INTO Clientes
        (Nome, CPF, Email, Telefone)
        VALUES (?,?,?,?)
        ''', parametro=(nome.upper(), cpf, email, telefone))

        return resultado_acao

    def cadastrar_produtos(self, descricao, cod_barras, preco, quantidade):
        resultado_acao = self.acao_sql('''
        INSERT INTO Produtos
        (Descricao, Cod_Barras, Preco, Quantidade)
        VALUES (?,?,?,?)
        ''', parametro=(descricao.upper(), cod_barras, preco, quantidade))

        return resultado_acao

    def cadastrar_usuarios(self, nome, cpf, funcao, nome_login, senha):
        resultado_acao = self.acao_sql('''
        INSERT INTO Usuarios
        (Nome, CPF, Funcao, Nome_Login, Senha)
        VALUES (?,?,?,?,?)
        ''', parametro=(nome.upper(), cpf, funcao.upper(), nome_login.upper(), senha))

        return resultado_acao

class Alterar(Acao_SQL):
    def alterar_pedidos(self):
        pass

    def alterar_itens_do_pedido(self): 
        pass

    def alterar_clientes(self, id_cliente, nome, cpf, email, telefone):
        self.acao_sql('''
        UPDATE Clientes
        SET Nome=?, CPF=?, Email=?, Telefone=?
        WHERE Id_Cliente = ?;
        ''', parametro=(nome, cpf, email, telefone, id_cliente))

    def alterar_produtos(self):
        pass

    def alterar_usuarios(self):
        pass

class Excluir(Acao_SQL):
    def excluir_pedidos(self, id_pedido):
        self.acao_sql('''
        DELETE FROM Pedidos
        WHERE Id_Pedido = ?;
        ''', parametro=(id_pedido,))

    def excluir_itens_do_pedido(self):
        pass

    def excluir_clientes(self, id_cliente):
        self.acao_sql('''
        DELETE FROM Clientes
        WHERE Id_Cliente = ?;
        ''', parametro=(id_cliente,))

    def excluir_produtos(self, id_produto):
        self.acao_sql('''
        DELETE FROM Produtos
        WHERE Id_Produto = ?;
        ''', parametro=(id_produto,))

    def excluir_usuarios(self, id_usuario):
        self.acao_sql('''
        DELETE FROM Usuarios
        WHERE Id_Usuario = ?;
        ''', 
        (id_usuario,))

class Exibir(Acao_SQL):
    def exibir_clientes(self):
        clientes = self.acao_sql('''
        SELECT Id_Cliente, Nome, CPF, Email, Telefone
        FROM Clientes
        ORDER BY Id_Cliente ASC
        ''', fetchall=True)

        lista_clientes = [list(cliente) for cliente in clientes]

        return lista_clientes
    
    def exibir_produtos(self):
        produtos = self.acao_sql('''
        SELECT Id_Produto, Descricao, Cod_Barras, Preco, Quantidade
        FROM Produtos
        ORDER BY Id_Produto ASC
        ''', fetchall=True)

        lista_produtos = [list(produto) for produto in produtos]

        return lista_produtos

    def exibir_pedidos(self):
        pedidos = self.acao_sql('''
        SELECT Id_Pedido, Nome_Cliente, Status, Valor_Pedido, Dt_Pedido, Dt_Alteracao
        FROM Pedidos
        ORDER BY Id_Pedido ASC
        ''', fetchall=True)

        lista_pedidos = [list(pedido) for pedido in pedidos]

        return lista_pedidos

    def exibir_itens_do_pedido(self):
        pass

    def exibir_usuarios(self):
        usuarios = self.acao_sql('''
        SELECT Id_Usuario, Nome, CPF, Funcao, Nome_Login, Senha
        FROM Usuarios
        ORDER BY Id_Usuario ASC
        ''', fetchall=True)

        lista_usuarios = [list(usuario) for usuario in usuarios]

        return lista_usuarios
    
class Exportar:
    pass

class Importar:
    pass




class Clientes:
    def exportar_clientes_excel(caminho_salvar_planilha):
        try:
            conexao = sqlite3.connect('Base de Dados.db')

            cursor = conexao.cursor()

            cursor.execute('''SELECT * FROM Clientes
            ORDER BY Nome ASC''')

            clientes_banco = cursor.fetchall()
            
            lista_clientes = [list(cliente) for cliente in clientes_banco]

            df = pd.DataFrame(lista_clientes, columns=['Nome','CPF','Telefone'])

            df.to_excel(caminho_salvar_planilha, index=False)

        finally:
            cursor.close()
            conexao.close()

    def importar_clientes_excel(caminho_planilha):
        planilha = openpyxl.load_workbook(caminho_planilha).active

        contador = 0
        erros = 0
        total_linhas = planilha.max_row - 1

        for linha in planilha.iter_rows(min_row=2):
            try:
                nome = linha[0].value
                cpf = linha[1].value
                telefone = linha[2].value

            except IndexError:
                yield False, False, False

            if nome is None or cpf is None or telefone is None:
                contador += 1
                erros += 1
            
            else:
                try:
                    conexao = sqlite3.connect('Base de Dados.db')

                    cursor = conexao.cursor()

                    cursor.execute('''INSERT INTO Clientes (Nome, CPF, Telefone) 
                    VALUES (?,?,?)''',
                    (nome.upper(), cpf, telefone))

                    conexao.commit()

                except sqlite3.IntegrityError:
                    cursor.execute('''
                    UPDATE Clientes
                    SET Nome = ?, CPF = ?, Telefone = ?
                    WHERE CPF = ?;''',
                    (nome.upper(), cpf, telefone, cpf))

                    conexao.commit()

                finally:
                    cursor.close()
                    conexao.close()
                
                contador += 1
            
            yield contador, total_linhas, erros


class Produtos:
    def exportar_produtos_excel(caminho_salvar_planilha):
        try:
            conexao = sqlite3.connect('Base de Dados.db')

            cursor = conexao.cursor()

            cursor.execute('''SELECT Descricao, Cod_Barras, Preco, Quantidade FROM Produtos
            ORDER BY ID_Produto ASC''')

            produtos_banco = cursor.fetchall()
            
            lista_produtos = [list(produto) for produto in produtos_banco]

            df = pd.DataFrame(lista_produtos, columns=['Descrição','Cod_Barras','Preço','Quantidade'])

            df.to_excel(caminho_salvar_planilha, index=False)

        finally:
            cursor.close()
            conexao.close()

    def importar_produtos_excel(caminho_planilha):
        planilha = openpyxl.load_workbook(caminho_planilha).active

        erros = 0
        contador = 0
        total_linhas = planilha.max_row - 1

        for linha in planilha.iter_rows(min_row=2):
            try:
                descricao = linha[0].value
                cod_barras = linha[1].value
                preco = linha[2].value
                quantidade = linha[3].value

            except IndexError:
                yield False, False, False

            if descricao is None or cod_barras is None or preco is None or quantidade is None:
                contador += 1
                erros += 1
                
            else:
                try:
                    conexao = sqlite3.connect('Base de Dados.db')

                    cursor = conexao.cursor()

                    cursor.execute('''INSERT INTO Produtos (Descricao, Cod_Barras, Preco, Quantidade) 
                    VALUES (?,?,?,?)''',
                    (descricao.upper(), cod_barras, preco, quantidade))

                    conexao.commit()

                except sqlite3.IntegrityError:
                    cursor.execute('''
                    UPDATE produtos
                    SET Descricao = ?, Cod_Barras = ?, Preco = ?, Quantidade = ?
                    WHERE Cod_Barras = ?;''',
                    (descricao, cod_barras, preco, quantidade, cod_barras))

                    conexao.commit()

                finally:
                    cursor.close()
                    conexao.close()
                
                contador += 1
            
            yield contador, total_linhas, erros


class Pedidos:
    def exportar_pedidos_excel(caminho_salvar_planilha):
        try:
            conexao = sqlite3.connect('Base de Dados.db')

            cursor = conexao.cursor()

            cursor.execute('''SELECT ID_Pedido, Nome_Cliente, CPF_Cliente, Valor, Status, Dt_Pedido, Dt_Alteracao, Produtos
            FROM Pedidos
            ORDER BY ID_Pedido ASC''')

            pedidos_banco = cursor.fetchall()
            
            lista_pedidos = [list(pedido) for pedido in pedidos_banco]

            df = pd.DataFrame(lista_pedidos, columns=['ID_PEDIDO','NOME_CLIENTE','CPF_CLIENTE','VALOR','STATUS','DATA_PEDIDO','DATA_ALTERAÇÃO','PRODUTOS'])

            df.to_excel(caminho_salvar_planilha, index=False)

        finally:
            cursor.close()
            conexao.close() 
